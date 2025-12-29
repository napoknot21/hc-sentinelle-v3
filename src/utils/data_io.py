from __future__ import annotations

import os
import time
import hashlib
import openpyxl
import xlwings as xw
import polars as pl

from typing import Dict, Optional, List, Tuple

from src.config.parameters import *
from src.config.paths import *
from src.utils.logger import *
from src.utils.formatters import numeric_cast_expr_from_utf8, date_cast_expr_from_utf8


def load_excel_to_dataframe (
        
        excel_file_abs_pth : str,
        sheet_name : str = "Sheet1",
        specific_cols : Optional[List] = None,
        schema_overrides : Optional[Dict] = None,
        cast_num : bool = True,
        allow_us_mdy: bool = False,        # set True if you truly expect mm/dd/yyyy
        date_formats: Optional[List[str]] = None,
        
    ) -> Tuple[Optional[pl.DataFrame], Optional[str]]  :
    """
    Loads an Excel file and returns it as a Polars dataframe, applying schema overrides.

    Args:
        excel_file_abs_pth (str): Absolute path to the Excel file.
        sheet_name (str): Sheet name to read from.
        specific_cols (list): List of columns to import. If None, then all columns
        schema_overrides (dict): Dictionary mapping column names to Polars types.

    Returns:
        df (pl.DataFrame | None) : Loaded dataframe, or None if an error occurs.
    """
    if not os.path.isfile(excel_file_abs_pth) :
        
        log(f"[-] File not found : {excel_file_abs_pth}", "error")
        return None, None

    if sheet_name is None or sheet_name == "" :
        sheet_name = 0 # The default sheet index

    try :

        # sving the time of execution
        start = time.time()
        
        df = pl.read_excel(

            source=excel_file_abs_pth,
            sheet_name=sheet_name,
            columns=specific_cols,              # If None, pl manage this case
            schema_overrides=None if cast_num is True else schema_overrides   # If None, Polars manages this case
        
        )

        if cast_num is False or schema_overrides is None :

            csv_bytes = df.write_csv().encode("utf-8")
            md5_hash = hashlib.md5(csv_bytes).hexdigest()

            log(f"[*] [POLARS] Read in {time.time() - start:.2f} seconds from {excel_file_abs_pth}", "info")
            
            return df, md5_hash


        actual = dict(zip(df.columns, df.dtypes))
        exprs: list[pl.Expr] = []

        for col, target_dtype in schema_overrides.items() :

            if col not in actual :
                continue  # column not present in this sheet

            actual_dtype = actual[col]
            
            # If matches already, skip
            if actual_dtype == target_dtype :
                continue

            # If the column is a Date (in string format)
            is_target_date = target_dtype in (pl.Date, pl.datetime)

            if is_target_date and actual_dtype == pl.Utf8 :

                exprs.append(

                    date_cast_expr_from_utf8(

                        col,
                        to_datetime=(target_dtype == pl.datetime),
                        formats=date_formats,
                        allow_us_mdy=allow_us_mdy,
                        enable_excel_serial=True

                    )

                )

                continue
                        
            # If target is numeric and actual is Utf8 → clean+cast from string
            is_target_float = target_dtype in (pl.Float32, pl.Float64)
            is_target_int = target_dtype in (pl.Int8, pl.Int16, pl.Int32, pl.Int64, pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64)
            
            if (is_target_float or is_target_int) and actual_dtype == pl.Utf8 :
                
                exprs.append(

                    numeric_cast_expr_from_utf8(
                        col,
                        target_dtype,
                        decimal=".",            # change to "," if your inputs are EU-style
                        int_rounding="nearest", # or "truncate"/"floor"/"ceil"
                    )

                )

                continue
            
            # Fallback: try a permissive cast (won’t crash on bad values)
            exprs.append(pl.col(col).cast(target_dtype, strict=False).alias(col))

        if exprs :
            df = df.with_columns(exprs)

        csv_bytes = df.write_csv().encode("utf-8")
        md5_hash = hashlib.md5(csv_bytes).hexdigest()

        log(f"[*] [POLARS] Read in {time.time() - start:.2f} seconds from {excel_file_abs_pth}", "info")

        return df, md5_hash
    
    except Exception as e :

        log(f"[-] Error while converting the Excel file : {e}", "error")
        
        return None, None


def load_csv_to_dataframe (
        
        csv_abs_path : str,
        specific_cols : Optional[List] = None,
        schema_overrides : Optional[Dict] = None
    
    ) -> Tuple[Optional[pl.DataFrame], Optional[str]] :
    """
    Loads a CSV file into a Polars DataFrame, with optional column filtering and type overrides.

    Args:
        csv_abs_path (str): Absolute path to the CSV file.
        specific_cols (list, optional): List of columns to read. If None, all columns are read.
        schema_overrides (dict, optional): Dictionary mapping column names to Polars data types.

    Returns:
        pl.DataFrame | None: The loaded Polars DataFrame, or None if an error occurred.
    """
    if not os.path.isfile(csv_abs_path) :

        log(f"\n[-] File not found : {csv_abs_path}\n", "error")
        return None, None
    
    try :

        start = time.time()

        df = pl.read_csv(

            source=csv_abs_path,
            columns=specific_cols,
            schema_overrides=schema_overrides,

            low_memory=False,   # In order to get speed lecture
            n_threads=4         # In order to get speed lecture

        )

        csv_bytes = df.write_csv().encode("utf-8")
        md5_hash = hashlib.md5(csv_bytes).hexdigest()

        log(f"[*] [POLARS] Read in {time.time() - start:.2f} seconds from CSV {csv_abs_path}", "info")

        return df, md5_hash
    
    except Exception as e :

        log(f"[-] Error during convertion of CSV {csv_abs_path} : {e}", "error")
    
        return None, None


def load_json_to_dataframe (
        
        json_abs_path : str,
        schema_overrides : Optional[Dict] = None
    
    ) -> Optional[pl.DataFrame]  :
    """
    Loads a JSON file into a Polars DataFrame.

    Args:
        json_abs_path (str): Absolute path to the JSON file.
        schema_overrides (dict, optional): Dictionary mapping column names to Polars data types.

    Returns:
        pl.DataFrame | None: The loaded Polars DataFrame, or None if an error occurred.
    """
    if not os.path.isfile(json_abs_path) :

        log(f"[-] File {json_abs_path} not found...", "error")
        return None, None


    try :
            
        start = time.time()

        df = pl.read_json(

            source=json_abs_path,
            schema_overrides=schema_overrides

        )

        csv_bytes = df.write_csv().encode("utf-8")
        md5_hash = hashlib.md5(csv_bytes).hexdigest()

        log(f"[*] [POLARS] Read in {time.time() - start:.2f} seconds from JSON {json_abs_path}", "info")

        return df, md5_hash

    except Exception as e :

        log(f"[-] Error during convertion of {json_abs_path} : {e}", "error")
    
        return None, None


def export_dataframe_to_excel (df : pl.DataFrame, sheet_name : str = "Sheet1", output_abs_path : str = None) -> Dict :
    """
    Exports a Polars DataFrame to an Excel file.

    Args:
        df (pl.DataFrame): The DataFrame to export.
        sheet_name (str): Name of the Excel sheet. Default is "Sheet1".
        output_abs_path (str): Absolute path to output the Excel file.

    Returns:
        response (dict) : Dictionary containing:
            - 'success' (bool): Whether the export was successful.
            - 'message' (str): Description of the result.
            - 'path' (str): Output file path if successful.
    """
    response = {

        'success' :  False,
        'message' : None,
        'path' : None

    }

    if output_abs_path is None or output_abs_path == "" :

        response["message"] = "Output path not specified."
        return response
    
    try :

        start = time.time()

        df.write_excel(
        
            workbook=output_abs_path,
            worksheet=sheet_name

        )

        duration = time.time() - start

        log(f"[+] [POLARS] Excel written in {duration:.2f} seconds to {output_abs_path}", "info")

        response["success"] = True
        response["message"] = "Export Successful"
        response["path"] = output_abs_path
    
    except Exception as e :

        log(f"[-] Failed to export DataFrame to Excel: {e}", "error")

        response["message"] = f"Export failed : {e}"
    
    finally :

        return response
    

def export_dataframe_to_json (df : pl.DataFrame, output_abs_path : Optional[str] = None) -> Dict :
    """
    Exports a Polars DataFrame to a JSON file.

    Args:
        df (pl.DataFrame): The Polars DataFrame to export.
        output_abs_path (str, optional): The absolute path where the JSON file will be saved. If None, the file is not saved.

    Returns:
        response (dict) : A dictionary containing:
            - 'success' (bool): Whether the export was successful or not.
            - 'message' (str): A message describing the result.
            - 'path' (str or None): The output file path if successful, None otherwise.
    """
    response = {

        'success' : False,
        'message' : None,
        'path' : None

    }

    if output_abs_path is None or output_abs_path == "" :
        
        response['message'] = "Output path not specified."
        return response
    
    try :

        start = time.time()

        df.write_json(

            file=output_abs_path

        )

        duration = time.time() - start
        log(f"[+] [POLARS] JSON written in {duration:.2f} seconds to {output_abs_path}", "info")

        response["success"] = True
        response["message"] = "Export Successful"
        response['path'] = output_abs_path

    except Exception as e :

        log(f"[-] Failed to export DataFrame to JSON: {e}", "error")

        response["message"] = f"Export failed : {e}"

    finally :

        return response
    

def export_dataframe_to_csv (df : pl.DataFrame, separator : str = ",", output_abs_path : Optional[str] = None) -> Dict :
    """
    Exports a Polars DataFrame to a CSV file.

    Args:
        df (pl.DataFrame): The Polars DataFrame to export.
        decimal_coma (bool, optional): Whether to use a comma as the decimal separator (True) or a period (False). Defaults to False.
        output_abs_path (str): The absolute path where the CSV file will be saved. If None, the file is not saved.

    Returns:
        response (dict) : A dictionary containing:
            - 'success' (bool): Whether the export was successful or not.
            - 'message' (str): A message describing the result.
            - 'path' (str or None): The output file path if successful, None otherwise.
    """
    response = {

        'success' : False,
        'message' : None,
        'path' : None

    }

    if output_abs_path is None or output_abs_path == "" :
        
        response['message'] = "Output path not specified."
        return response
    
    try : 

        start = time.time()

        df.write_csv(

            file=output_abs_path,
            # decimal_comma=decimal_comma # Error here due to the polars version
            separator=separator
        )

        duration = time.time() - start
        log(f"[+] [POLARS] CSV written in {duration:.2f} seconds to {output_abs_path}", "info")

        response["success"] = True
        response["message"] = "Export Successful"
        response['path'] = output_abs_path

    except Exception as e :

        log(f"[-] Failed to export DataFrame to CSV: {e}", "error")

        response["message"] = f"Export failed : {e}"

    finally :

        return response
    

def export_excel_to_pdf (file_abs_path : Optional[str] = None, output_filename : Optional[str] = None, output_dir_path : Optional[str] = None) :
    """
    
    """
    response = {

        "success" : False,
        "message" : None,
        "path" : None

    }

    if not os.path.exists(file_abs_path) :

        # No file existed. Error returned
        return response
    

    output_dir_path = PAYMENTS_FILES_ABS_PATH if output_dir_path is None else output_dir_path
    os.makedirs(output_dir_path, exist_ok=True)
    
    with xw.App() as app :

        # user will not even see the excel opening up
        app.visible = False
        
        book = app.books.open(file_abs_path)
        sheet = book.sheets[0]
       
        # Save excel workbook as pdf
        full_path = os.path.join(output_dir_path, output_filename)
        sheet.to_pdf(path=full_path, show=False)
        
        response["success"] = True
        response["path"] = full_path

    return response


def convert_payement_to_excel (
        
        payment : Optional[Tuple] = None,

        template_abs_path : Optional[str] = None,
        dir_abs_path : Optional[str] = None,
        columns_index : Optional[List] = None

    ) :

    """
    
    """
    if payment is None or len(payment) <= 0 :
        return None
    
    columns_index =  PAYMENTS_EXCEL_COLUMNS if columns_index is None else columns_index
    template_abs_path = PAYMENTS_EXCEL_TEMPLATE_ABS_PATH if template_abs_path is None else template_abs_path
    dir_abs_path = PAYMENTS_FILES_ABS_PATH if dir_abs_path is None else dir_abs_path

    os.makedirs(dir_abs_path, exist_ok=True)

    workbook = openpyxl.load_workbook(template_abs_path)
    sheet = workbook.active

    row_idx = 3

    for value, col_letter in zip(payment, columns_index) :

        cell_ref = f"{col_letter}{row_idx}"
        cell = sheet[cell_ref]
        
        cell.value = value

        if isinstance(value, (dt.date, dt.datetime)):
            cell.number_format = "DD/MM/YYYY"

    filename = f"Payment_instructions_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx" 
    filled_path = os.path.join(dir_abs_path, filename)
        
    workbook.save(filled_path)
    
    return filled_path


def convert_ubs_instruction_payments_to_excel (
        
        payments : Optional[Tuple] = None,
        template_abs_path : Optional[str] = None,

        filename : Optional[str] = None,
        dir_abs_path : Optional[str] = None,

        columns_index : Optional[List[str]] = None,

    ) :
    """
    Docstring for convert_ubs_instruction_payment_to_excel
    """
    response = {

        "success" : False,
        "message" : None,
        "path" : None

    }

    if payments is None or len(payments) <= 0 :

        log("[-] No collateral passed for conversion.")
        return response
    
    template_abs_path = UBS_PAYMENTS_INTRUCTION_TEMPLATE_ABS_PATH if template_abs_path is None else template_abs_path 
    
    filename = f"UBS_Payment_Instruction_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx" if filename is None else filename
    dir_abs_path = PAYMENTS_FILES_ABS_PATH if dir_abs_path is None else dir_abs_path
    columns_index = UBS_PAYMENTS_EXCEL_COLUMNS if columns_index is None else columns_index

    os.makedirs(dir_abs_path, exist_ok=True)

    workbook = openpyxl.load_workbook(template_abs_path)
    sheet = workbook.active

    row_idx = 7
    line_reason = 2

    for payment in payments :

        for value, col_letter in zip(payment, columns_index) :

            cell_ref = f"{col_letter}{row_idx}"
            cell = sheet[cell_ref]
            
            if value is None :
                continue

            cell.value = value

            if isinstance(value, (dt.date, dt.datetime)):
                cell.number_format = "DD/MM/YYYY"

        row_idx += line_reason

    filled_path = os.path.join(dir_abs_path, filename)
    
    try :

        workbook.save(filled_path)

        response["success"] = True
        response["path"] = filled_path
        response["message"] = "Cash Transfer For billateral OTC Successfully loaded and filled"

        log(f"[+] {response["message"]}")

    except : 
        log("[-] Error during saving process.")
    
    return response


def convert_ubs_collateral_management_to_excel (
        
        collaterals : Optional[Tuple] = None,
        template_abs_path : Optional[str] = None,

        filename : Optional[str] = None,
        dir_abs_path : Optional[str] = None,
    
    ) :
    """
    Docstring for convert_ubs_collateral_management_to_excel
    
    :param collaterals: Description
    :type collaterals: Optional[Tuple]
    :param template_abs_path: Description
    :type template_abs_path: Optional[str]
    :param dir_abs_path: Description
    :type dir_abs_path: Optional[str]
    """
    response = {

        "success" : False,
        "message" : None,
        "path" : None

    }

    if collaterals is None or len(collaterals) <= 0 :

        log("[-] No collateral passed for conversion.")
        return response
    
    template_abs_path = UBS_COLLATERAL_MANAGEMENT_TEMPLATE_ABS_PATH if template_abs_path is None else template_abs_path 
    
    filename = f"UBS_Collateral_Management_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx" if filename is None else filename
    dir_abs_path = PAYMENTS_FILES_ABS_PATH if dir_abs_path is None else dir_abs_path

    os.makedirs(dir_abs_path, exist_ok=True)

    workbook = openpyxl.load_workbook(template_abs_path)
    sheet = workbook.active

    row_idx = 7
    line_reason = 2

    for collateral in collaterals :
        print("Hello")

    filled_path = os.path.join(dir_abs_path, filename)
    
    try :

        workbook.save(filled_path)

        response["success"] = True
        response["path"] = filled_path
        response["message"] = "Collateral Successfully loaded and filled"

        log(f"[+] {response["message"]}")

    except : 
        log("[-] Error during saving process.")
    
    return response

